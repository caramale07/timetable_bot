import telebot as tl
from openpyxl import Workbook,load_workbook

from calendar import week
import datetime as dt
from tokenize import group
import xml.etree.ElementTree as ET
from calendar import week

'''---------------------------------------------------------'''


class Card:

    def __init__(self, lessonid: str, classroomids: str, period: int, weeks: str, days: str) -> None:
        self.lessonid = lessonid
        self.classroomids = classroomids
        self.period = period
        self.weeknum = weeks
        self.daynum = days
        self.faculty_ids = ""
        self.faculty = []
        self.group_name = []
        self.name = ""
        self.room = ""
        self.week = ""
        self.day = ""
        self.teacher_name = ""

class Lesson:
    
    def __init__(self,id: str, subjectid: str,teacherids:str,groupids:str) -> None:
        self.id = id
        self.subjectid = subjectid
        self.teacherids = teacherids
        self.groupids = groupids

class Teacher:


    def __init__(self, id: str, name: str) -> None:
        self.id = id
        self.name = name

class Subject:

    def __init__(self, id: str, name: str) -> None:
        self.id = id
        self.name = name

class Classroom:

    def __init__(self, id: str, name: int) -> None:
        self.id = id
        self.name = name

class Weeksdef:


    def __init__(self, id: str, name: str, short: int, weeks: str) -> None:
        self.id = id
        self.name = name
        self.short = short
        self.weeks = weeks

class Daysdef:


    def __init__(self,id: str, name: str, days: str ) -> None:
        self.id = id
        self.name = name
        self.days = days

class Group:

    def __init__(self, id: str, name: str, classid: str) -> None:
        self.id = id
        self.name = name
        self.classid = classid

class Class:

    def __init__(self, id: str, name: str) -> None:
        self.id = id
        self.name = name


way =  r"D:\CODE\study project\ufaz_timetable_bot\timetable.xml"


def initialization(way):
    global root
    tree = ET.parse(way)
    root = tree.getroot()

    """Parser"""

    def xml_parser(first_tag,second_tag,list_of_inside_tags): # It works, DO NOT TOUCH!!! 
        full_list = []
        for main in root.findall(first_tag):
            for secondary in main.findall(second_tag):
                l = []
                for tag in list_of_inside_tags:
                    txt = secondary.get(tag)
                    l.append(txt)
                full_list.append(l)
        return full_list

    def starting_programm():
        """Info Export"""  

        cards = (xml_parser("cards","card",["lessonid","classroomids","period","weeks","days"]))
        global list_of_cards
        list_of_cards = []

        for i in range(len(cards)):
            try:
                list_of_cards.append(Card(cards[i][0],cards[i][1],cards[i][2],cards[i][3],cards[i][4]))
            except:
                pass


        lessons = (xml_parser("lessons","lesson",["id","subjectid","teacherids","groupids"]))
        list_of_lessons = []

        for i in range(len(lessons)):
            try:
                list_of_lessons.append(Lesson(lessons[i][0],lessons[i][1],lessons[i][2],lessons[i][3]))
            except:
                pass

        subjects = xml_parser("subjects","subject",["id","name"])
        list_of_subjects = []

        for i in range(len(subjects)):
            try:
                list_of_subjects.append(Subject(subjects[i][0],subjects[i][1]))
            except:
                pass

        classrooms = xml_parser("classrooms","classroom",["id","name"])
        list_of_classrooms = []

        for i in range(len(classrooms)):
            try:
                list_of_classrooms.append(Classroom(classrooms[i][0],classrooms[i][1]))
            except:
                pass

        weeksdefs = xml_parser("weeksdefs","weeksdef",["id","name","short","weeks"])
        list_of_weeksdefs = []

        for i in range(len(weeksdefs)):
            try:
                int(weeksdefs[i][2])  #чтобы не получить варианы any day , every day 
                list_of_weeksdefs.append(Weeksdef(weeksdefs[i][0],weeksdefs[i][1],weeksdefs[i][2],weeksdefs[i][3]))
            except:
                pass

        daydefs = xml_parser("daysdefs","daysdef",["id","name","days"])
        list_of_daydefs = []

        for i in range(len(daydefs)):
            try:
                list_of_daydefs.append(Daysdef(daydefs[i][0],daydefs[i][1],daydefs[i][2]))
            except:
                pass

        teachers = xml_parser("teachers","teacher",["id","name"])
        list_of_teachers = []

        for i in range(len(teachers)):
            try:
                list_of_teachers.append(Teacher(teachers[i][0],teachers[i][1]))
            except:
                pass

        groups = (xml_parser("groups","group",["id","name","classid"]))
        list_of_groups = []

        for i in range(len(groups)):
            try:
                list_of_groups.append(Group(groups[i][0],groups[i][1],groups[i][2]))
            except:
                pass

        classes = (xml_parser("classes","class",["id","name"]))
        list_of_classes = []

        for i in range(len(classes)):
            try:
                list_of_classes.append(Class(classes[i][0],classes[i][1]))
            except:
                pass


        """ Transformation of Information """

        for i in range(len(list_of_cards)):
            for j in range(len(list_of_lessons)):
                if list_of_cards[i].lessonid == list_of_lessons[j].id:
                    list_of_cards[i].name = list_of_lessons[j].subjectid


        for i in range(len(list_of_cards)):
            for j in range(len(list_of_lessons)):
                if list_of_cards[i].lessonid == list_of_lessons[j].id:
                    list_of_cards[i].teacher_name = list_of_lessons[j].teacherids


        for i in range(len(list_of_cards)):
            for j in range(len(list_of_teachers)):
                if list_of_cards[i].teacher_name == list_of_teachers[j].id:
                    list_of_cards[i].teacher_name = list_of_teachers[j].name



        for i in range(len(list_of_cards)):
            for j in range(len(list_of_subjects)):
                if list_of_cards[i].name == list_of_subjects[j].id:
                    list_of_cards[i].name = list_of_subjects[j].name


        for i in range(len(list_of_cards)):
            for j in range(len(list_of_classrooms)):
                if list_of_cards[i].classroomids == list_of_classrooms[j].id:
                    list_of_cards[i].room = list_of_classrooms[j].name


        for i in range(len(list_of_cards)):
            for j in range(len(list_of_weeksdefs)):
                if list_of_cards[i].weeknum == list_of_weeksdefs[j].weeks:
                    list_of_cards[i].week = list_of_weeksdefs[j].name


        for i in range(len(list_of_cards)):
            for j in range(len(list_of_daydefs)):
                if list_of_cards[i].daynum == list_of_daydefs[j].days:
                    list_of_cards[i].day = list_of_daydefs[j].name



        """Ёбка with groups"""


        for card in list_of_cards:
            for lesson in list_of_lessons:
                if card.lessonid == lesson.id:
                    group_ids_temporal = (lesson.groupids).split(',')
                    for group_id in group_ids_temporal:
                        for group in list_of_groups:
                            if group_id == group.id:
                                card.group_name.append(group.name)
                                for cls in list_of_classes:
                                    if group.classid == cls.id:
                                        card.faculty.append(cls.name)

        """Entire class to groups"""


        for card in list_of_cards:
            if "Entire class" in card.group_name:
                card.group_name.append("Group 1")
                card.group_name.append("Group 2")


    starting_programm()

initialization(way)

day_true = {
    "Monday":1,
    "Tuesday":2,
    "Wednesday":3,
    "Thursday":4,
    "Friday":5
}

week_true={
    "Week 1 (24.01-28.01)" :1,
    "Week 2 (31.01-4.02)"  :2,
    "Week 3 (07.02-11.02)" :3,
    "Week 4 (14.02-18.02)" :4,
    "Week 5 (21.02-25.02)" :5,
    "Week 6 (28.02-04.03)" :6,
    "Week 7 (07.03-11.03)" :7,
    "Week 8 (14.03-18.03)" :8,
    "Week 9 (28.03-01.04)" :9,
    "Week 10 (04.04-08.04)" :10,
    "Week 11 (11.04-15.04)" :11,
    "Week 12 (18.04-22.04)" :12,
    "Week 13 (25.04-29.04)" :13,
    "Week 14 (02.05-06.05)" :14,
    "Week 15 (09.05-13.05)" :15,
    "Week 16 (16.05-20.05)" :16,
    "Revision Week 17 (23.05-27.05)" :17,
    "Exam Week 18 (30.05-03.05)" :18,
    "Exam Week 19 (06.06-10.06)" :19,
    "Exam Week 20 (13.06-17.06)" :20
}


def find():
        for i in list_of_cards:
            if "A" in i.faculty:
                if i.group_name == "Group 2":
                    print(i.name + "___" +i.teacher_name)


# find()
def check(a):
    try:
        # print(list_of_cards[a].lessonid)
        print(a.name)
        # print(list_of_cards[a].classroomids)
        print(a.room)
        print("#",a.period)
        # print(list_of_cards[a].weeknum)
        print(a.week)
        # print(list_of_cards[a].daynum)
        print(a.day)
        print(a.teacher_name)
        print(a.faculty)
        print(a.group_name)
    except:
        pass


# check(3366)



now = dt.datetime.now()
current_time = now.strftime("%H:%M") # 22:59

todays_date = dt.date.today() # 2022-03-16
current_day = todays_date.day # 16
current_month = todays_date.month # 3
current_week_day = dt.datetime.today().isoweekday() # 3

'''tomorrow'''

tomorrows_date = dt.date.today() + dt.timedelta(days=1)
tomorrow_day = tomorrows_date.day
tomorrow_month = tomorrows_date.month
tomorrow_week_day = tomorrows_date.isoweekday()


def get_week_number(day,month):    
    if month == 3:
        if day>=14 and day<=20:
            return 8
        elif day>=28:
            return 9
    elif month == 4:
        if day>=1 and day<=3:
            return 9
        if day>=4 and day<=10:
            return 10
        if day>=11 and day<=17:
            return 11
        if day>=18 and day<=24:
            return 12
        if day>=25:
            return 13
    elif month == 5:
        if day == 1:
            return 13
        if day>=2 and day<=8:
            return 14
        if day>=9 and day<=15:
            return 15
        if day>=16 and day<=22:
            return 16
        if day>=23 and day<=29:
            return 17
        if day >=30:
            return 18
    elif month == 6:
        if day>=1 and day<=5:
            return 18
        if day>=6 and day<=12:
            return 19
        if day>=13 and day<=19:
            return 20
    else:
        return None
    
global current_week_number
current_week_number = get_week_number(current_day,current_month)

'''FIND LESSSONS'''


def is_time_between(begin_time, end_time, check_time=None):
    # If check time is not given, default to current UTC time
    check_time = check_time or dt.datetime.utcnow().time()

    if begin_time < end_time:
        return check_time >= begin_time and check_time <= end_time
    else: # crosses midnight
        return check_time >= begin_time or check_time <= end_time


ctime = dt.datetime.now().time() 
# ctime and current_time are the same things but 
# in different formats

'''current lesson period'''

def get_lesson_period(ctime):
    if is_time_between(dt.time(8,30),dt.time(10,00),ctime):
        return 1
    elif is_time_between(dt.time(10,00),dt.time(10,15),ctime):
        return 2
    elif is_time_between(dt.time(10,15),dt.time(11,45),ctime):
        return 2
    elif is_time_between(dt.time(11,45),dt.time(12,45),ctime):
        return 3
    elif is_time_between(dt.time(12,45),dt.time(14,15),ctime):
        return 3
    elif is_time_between(dt.time(14,15),dt.time(14,30),ctime):
        return 4
    elif is_time_between(dt.time(14,30),dt.time(16,00),ctime):
        return 4

# print(get_lesson_period(dt.time(9,00)))
# print(get_lesson_period(ctime))
 


'''MAIN FUNCTIONS'''

def _current_index(current_day,current_week_number,student_faculty,student_group):
    period  = get_lesson_period(ctime)
    week = current_week_number
    day = current_day
    faculty = student_faculty
    group = student_group
    for i in list_of_cards:
        if faculty in i.faculty:
            if group in i.group_name:
                if week == week_true[i.week]:
                    if day == day_true[i.day]:
                        if period == int(i.period):
                            return i

    return "N"

def _next_index(current_day,current_week_number,student_faculty,student_group):
    period  = get_lesson_period(ctime)
    week = current_week_number
    day = current_day
    faculty = student_faculty
    group = student_group
    try:
        for j in range(1,7):
            for i in list_of_cards:
                if faculty in i.faculty:
                    if group in i.group_name:
                        if week == week_true[i.week]:
                            if day == day_true[i.day]:
                                if period+j == int(i.period):
                                    return i
        return "N"
    except:
        return "N"

def _today_index(current_day,current_week_number,student_faculty,student_group):
    week = current_week_number
    day = current_day
    faculty = student_faculty
    group = student_group
    index_list = []
    for i in list_of_cards:
        # print(faculty)
        # print(i.faculty)
        if faculty in i.faculty:
            # print(group)
            # print(i.group_name)
            if group in i.group_name:
                # print(week)
                # print(week_true[i.week])
                if week == week_true[i.week]:
                    # print(day+1)
                    # print(day_true[i.day])
                    if day == day_true[i.day]:
                        index_list.append(i)

    index_list_orderd = ["l"]*7 # извиняюсь , знаю что надо преределать, но лень
    for i in index_list:
        p = int(i.period)-1
        index_list_orderd.pop(p)
        index_list_orderd.insert(p,i)
    if index_list:   
        return index_list_orderd
    else:
        return "N"

def _tomorrow_index(current_day,current_week_number,student_faculty,student_group):
    week = current_week_number
    day = current_day
    faculty = student_faculty
    group = student_group
    index_list = []
    for i in list_of_cards:
        # print(faculty)
        # print(i.faculty)
        if faculty in i.faculty:
            # print(group)
            # print(i.group_name)
            if group in i.group_name:
                # print(week)
                # print(week_true[i.week])
                if week == week_true[i.week]:
                    # print(day+1)
                    # print(day_true[i.day])
                    if day+1 == day_true[i.day]:
                        index_list.append(i)

    index_list_orderd = ["l"]*7 # извиняюсь , знаю что надо преределать, но лень
    for i in index_list:
        p = int(i.period)-1
        index_list_orderd.pop(p)
        index_list_orderd.insert(p,i)
    if index_list:   
        return index_list_orderd
    else:
        return "N"







'''-------------------------------------------------------------'''


'''data and support functions'''

path = r'D:\CODE\study project\ufaz_timetable_bot\groups.xlsx'
wb = load_workbook(path)
sheet = wb.active   

def update_sheet(student_group,id,path):
    wb=load_workbook(path)
    sheet = wb.active
    if student_group == 'A 1':
        
        for col in sheet.iter_cols(min_row=2, min_col=1, max_col=1,max_row=100):
            for cell in col:
                if cell.value==None:
                    cell.value=id
                    break
    if student_group == 'A 2':
        
        for col in sheet.iter_cols(min_row=2, min_col=2, max_col=2,max_row=100):
            for cell in col:
                if cell.value==None:
                    cell.value=id
                    break
    if student_group == 'B 1':
        
        for col in sheet.iter_cols(min_row=2, min_col=3, max_col=3,max_row=100):
            for cell in col:
                if cell.value==None:
                    cell.value=id
                    break
    if student_group == 'B 2':
        
        for col in sheet.iter_cols(min_row=2, min_col=4, max_col=4,max_row=100):
            for cell in col:
                if cell.value==None:
                    cell.value=id
                    break
    if student_group == 'C 1':
        
        for col in sheet.iter_cols(min_row=2, min_col=5, max_col=5,max_row=100):
            for cell in col:
                if cell.value==None:
                    cell.value=id
                    break
    if student_group == 'C 2':
        
        for col in sheet.iter_cols(min_row=2, min_col=6, max_col=6,max_row=100):
            for cell in col:
                if cell.value==None:
                    cell.value=id
                    break
    if student_group == 'D 1':
        
        for col in sheet.iter_cols(min_row=2, min_col=7, max_col=7,max_row=100):
            for cell in col:
                if cell.value==None:
                    cell.value=id
                    break
    if student_group == 'D 2':
        
        for col in sheet.iter_cols(min_row=2, min_col=8, max_col=8,max_row=100):
            for cell in col:
                if cell.value==None:
                    cell.value=id
                    break
    if student_group == 'CE-20 1':
        
        for col in sheet.iter_cols(min_row=2, min_col=9, max_col=9,max_row=100):
            for cell in col:
                if cell.value==None:
                    cell.value=id
                    break
    if student_group == 'CE-20 2':
        
        for col in sheet.iter_cols(min_row=2, min_col=10, max_col=10,max_row=100):
            for cell in col:
                if cell.value==None:
                    cell.value=id
                    break
    if student_group == 'CS-20 1':
        
        for col in sheet.iter_cols(min_row=2, min_col=11, max_col=11,max_row=100):
            for cell in col:
                if cell.value==None:
                    cell.value=id
                    break
    if student_group == 'CS-20 2':
        
        for col in sheet.iter_cols(min_row=2, min_col=12, max_col=12,max_row=100):
            for cell in col:
                if cell.value==None:
                    cell.value=id
                    break
    if student_group == 'GE-20 1':
        
        for col in sheet.iter_cols(min_row=2, min_col=13, max_col=13,max_row=100):
            for cell in col:
                if cell.value==None:
                    cell.value=id
                    break
    if student_group == 'GE-20 2':
        
        for col in sheet.iter_cols(min_row=2, min_col=14, max_col=14,max_row=100):
            for cell in col:
                if cell.value==None:
                    cell.value=id
                    break
    if student_group == 'PE-20 1':
        
        for col in sheet.iter_cols(min_row=2, min_col=15, max_col=15,max_row=100):
            for cell in col:
                if cell.value==None:
                    cell.value=id
                    break
    if student_group == 'PE-20 2':
        
        for col in sheet.iter_cols(min_row=2, min_col=16, max_col=16,max_row=100):
            for cell in col:
                if cell.value==None:
                    cell.value=id
                    break
    if student_group == 'CE-19 1':
        
        for col in sheet.iter_cols(min_row=2, min_col=17, max_col=17,max_row=100):
            for cell in col:
                if cell.value==None:
                    cell.value=id
                    break
    if student_group == 'CE-19 2':
        
        for col in sheet.iter_cols(min_row=2, min_col=18, max_col=18, max_row=100):
            for cell in col:
                if cell.value==None:
                    cell.value=id
                    break
    if student_group == 'CS-19 1':
        
        for col in sheet.iter_cols(min_row=2, min_col=19, max_col=19, max_row=100):
            for cell in col:
                if cell.value==None:
                    cell.value=id
                    break
    if student_group == 'CS-19 2':
        
        for col in sheet.iter_cols(min_row=2, min_col=20, max_col=20, max_row=100):
            for cell in col:
                if cell.value==None:
                    cell.value=id
                    break
    if student_group == 'GE-19 1':
        
        for col in sheet.iter_cols(min_row=2, min_col=21, max_col=21, max_row=100):
            for cell in col:
                if cell.value==None:
                    cell.value=id
                    break
    if student_group == 'GE-19 2':
        
        for col in sheet.iter_cols(min_row=2, min_col=22, max_col=22, max_row=100):
            for cell in col:
                if cell.value==None:
                    cell.value=id
                    break
    if student_group == 'PE-19 1':
        
        for col in sheet.iter_cols(min_row=2, min_col=23, max_col=23, max_row=100):
            for cell in col:
                if cell.value==None:
                    cell.value=id
                    break
    if student_group == 'PE-19 2':
        
        for col in sheet.iter_cols(min_row=2, min_col=24, max_col=24, max_row=100):
            for cell in col:
                if cell.value==None:
                    cell.value=id
                    break
    if student_group == 'CE-18 1':
        
        for col in sheet.iter_cols(min_row=2, min_col=25, max_col=25, max_row=100):
            for cell in col:
                if cell.value==None:
                    cell.value=id
                    break
    if student_group == 'CE-18 2':
        
        for col in sheet.iter_cols(min_row=2, min_col=26, max_col=26, max_row=100):
            for cell in col:
                if cell.value==None:
                    cell.value=id
                    break
    if student_group == 'CS-18 1':
        
        for col in sheet.iter_cols(min_row=2, min_col=27, max_col=27, max_row=100):
            for cell in col:
                if cell.value==None:
                    cell.value=id
                    break
    if student_group == 'CS-18 2':
        
        for col in sheet.iter_cols(min_row=2, min_col=28, max_col=28, max_row=100):
            for cell in col:
                if cell.value==None:
                    cell.value=id
                    break
    if student_group == 'GE-18 1':
        
        for col in sheet.iter_cols(min_row=2, min_col=29, max_col=29, max_row=100):
            for cell in col:
                if cell.value==None:
                    cell.value=id
                    break
    if student_group == 'GE-18 2':
        
        for col in sheet.iter_cols(min_row=2, min_col=30, max_col=30, max_row=100):
            for cell in col:
                if cell.value==None:
                    cell.value=id
                    break
    if student_group == 'PE-18 1':
        
        for col in sheet.iter_cols(min_row=2, min_col=31, max_col=31, max_row=100):
            for cell in col:
                if cell.value==None:
                    cell.value=id
                    break
    if student_group == 'PE-18 2':
        
        for col in sheet.iter_cols(min_row=2, min_col=32, max_col=32, max_row=100):
            for cell in col:
                if cell.value==None:
                    cell.value=id
                    break
    if student_group == 'M1 DSAI-21 1':

        for col in sheet.iter_cols(min_row=2, min_col=33, max_col=33, max_row=100):
            for cell in col:
                if cell.value==None:
                    cell.value=id
                    break
    if student_group == 'M1 DSAI-21 2':
        
        for col in sheet.iter_cols(min_row=2, min_col=34, max_col=34, max_row=100):
            for cell in col:
                if cell.value==None:
                    cell.value=id
                    break
    if student_group == 'M1 GE-21 1':
        
        for col in sheet.iter_cols(min_row=2, min_col=35, max_col=36, max_row=100):
            for cell in col:
                if cell.value==None:
                    cell.value=id
                    break
    if student_group == 'M1 GE-21 2':
        
        for col in sheet.iter_cols(min_row=2, min_col=36, max_col=36, max_row=100):
            for cell in col:
                if cell.value==None:
                    cell.value=id
                    break
    if student_group == 'M1 CE-21 1':
        
        for col in sheet.iter_cols(min_row=2, min_col=37, max_col=37, max_row=100):
            for cell in col:
                if cell.value==None:
                    cell.value=id
                    break
    if student_group == 'M1 CE-21 2':
        
        for col in sheet.iter_cols(min_row=2, min_col=38, max_col=38, max_row=100):
            for cell in col:
                if cell.value==None:
                    cell.value=id
                    break
    wb.save(path)

def get_groups_dict_from(path):
    wb = load_workbook(path)
    sheet = wb.active
    groups_dict={
        "A 1": [],
        "A 2": [],
        "B 1": [],
        "B 2": [],
        "C 1": [],
        "C 2": [],
        "D 1": [],
        "D 2": [],
        "CE-20 1": [],
        "CE-20 2": [],
        "CS-20 1": [],
        "CS-20 2": [],
        "GE-20 1": [],
        "GE-20 2": [],
        "PE-20 1": [],
        "PE-20 2": [],# L2
        "CE-19 1": [],
        "CE-19 2": [],
        "CS-19 1": [],
        "CS-19 2": [],
        "GE-19 1": [],
        "GE-19 2": [],
        "PE-19 1": [],
        "PE-19 2": [],# L3
        "CE-18 1": [],
        "CE-18 2": [],
        "CS-18 1": [],
        "CS-18 2": [],
        "GE-18 1": [],
        "GE-18 2": [],
        "PE-18 1": [],
        "PE-18 2": [],# M
        "M1 DSAI-21 1": [],
        "M1 DSAI-21 2": [],
        "M1 GE-21 1": [],
        "M1 GE-21 2": [],
        "M1 CE-21 1": [],
        "M1 CE-21 2": []
    }
    for id in sheet['A']:
        if id.value!=None:
            groups_dict['A 1'].append(id.value)
    for id in sheet['B']:
        if id.value!=None:
            groups_dict['A 2'].append(id.value)
    for id in sheet['C']:
        if id.value!=None:
            groups_dict['B 1'].append(id.value)
    for id in sheet['D']:
        if id.value!=None:
            groups_dict['B 2'].append(id.value)
    for id in sheet['E']:
        if id.value!=None:
            groups_dict['C 1'].append(id.value)
    for id in sheet['F']:
        if id.value!=None:
            groups_dict['C 2'].append(id.value)
    for id in sheet['G']:
        if id.value!=None:
            groups_dict['D 1'].append(id.value)
    for id in sheet['H']:
        if id.value!=None:
            groups_dict['D 2'].append(id.value)
    for id in sheet['I']:
        if id.value!=None:
            groups_dict['CE-20 1'].append(id.value)
    for id in sheet['J']:
        if id.value!=None:
            groups_dict['CE-20 2'].append(id.value)
    for id in sheet['K']:
        if id.value!=None:
            groups_dict['CS-20 1'].append(id.value)
    for id in sheet['L']:
        if id.value!=None:
            groups_dict['CS-20 2'].append(id.value)
    for id in sheet['M']:
        if id.value!=None:
            groups_dict['GE-20 1'].append(id.value)
    for id in sheet['N']:
        if id.value!=None:
            groups_dict['GE-20 2'].append(id.value)
    for id in sheet['O']:
        if id.value!=None:
            groups_dict['PE-20 1'].append(id.value)
    for id in sheet['P']:
        if id.value!=None:
            groups_dict['PE-20 2'].append(id.value)
    for id in sheet['Q']:
        if id.value!=None:
            groups_dict['CE-19 1'].append(id.value)
    for id in sheet['R']:
        if id.value!=None:
            groups_dict['CE-19 2'].append(id.value)
    for id in sheet['S']:
        if id.value!=None:
            groups_dict['CS-19 1'].append(id.value)
    for id in sheet['T']:
        if id.value!=None:
            groups_dict['CS-19 2'].append(id.value)
    for id in sheet['U']:
        if id.value!=None:
            groups_dict['GE-19 1'].append(id.value)
    for id in sheet['V']:
        if id.value!=None:
            groups_dict['GE-19 2'].append(id.value)
    for id in sheet['W']:
        if id.value!=None:
            groups_dict['PE-19 1'].append(id.value)
    for id in sheet['X']:
        if id.value!=None:
            groups_dict['PE-19 2'].append(id.value)
    for id in sheet['Y']:
        if id.value!=None:
            groups_dict['CE-18 1'].append(id.value)
    for id in sheet['Z']:
        if id.value!=None:
            groups_dict['CE-18 2'].append(id.value)
    for id in sheet['AA']:
        if id.value!=None:
            groups_dict['CS-18 1'].append(id.value)
    for id in sheet['AB']:
        if id.value!=None:
            groups_dict['CS-18 2'].append(id.value)
    for id in sheet['AC']:
        if id.value!=None:
            groups_dict['GE-18 1'].append(id.value)
    for id in sheet['AD']:
        if id.value!=None:
            groups_dict['GE-18 2'].append(id.value)
    for id in sheet['AE']:
        if id.value!=None:
            groups_dict['PE-18 1'].append(id.value)
    for id in sheet['AF']:
        if id.value!=None:
            groups_dict['PE-18 2'].append(id.value)
    for id in sheet['AG']:
        if id.value!=None:
            groups_dict['M1 DSAI-21 1'].append(id.value)
    for id in sheet['AH']:
        if id.value!=None:
            groups_dict['M1 DSAI-21 2'].append(id.value)
    for id in sheet['AI']:
        if id.value!=None:
            groups_dict['M1 GE-21 1'].append(id.value)
    for id in sheet['AJ']:
        if id.value!=None:
            groups_dict['M1 GE-21 2'].append(id.value)
    for id in sheet['AK']:
        if id.value!=None:
            groups_dict['M1 CE-21 1'].append(id.value)
    for id in sheet['AL']:
        if id.value!=None:
            groups_dict['M1 CE-21 2'].append(id.value)
    
    return groups_dict

def get_student_group(id):
    student_group=''
    global path
    groups = get_groups_dict_from(path)
    for list_of_ids in list(groups.values()):
        if id in list_of_ids:
            student_group=list(groups.keys())[list(groups.values()).index(list_of_ids)]
    return student_group

token='***'
bot = tl.TeleBot(token, parse_mode=None)


'''START COMMAND if user's ID is not in database, it will be added'''

@bot.message_handler(commands=['start'])
def start(message):
    id = str(message.chat.id)
    student_group=get_student_group(id)
    if student_group=='':
        bot.send_message(id,"Firstly, /setgroup")
    else:
        bot.send_message(id,f"Bonjour, your group is {student_group}\n\n/setgroup - if you want to change the group")


'''SETGROUP command to set or change the group '''

@bot.message_handler(commands=['setgroup'])
def define_degree(message):
    id = str(message.chat.id)

    global path
    wb = load_workbook(path)
    sheet = wb.active

    for row in sheet.rows:
        for cell in row:
            if cell.value == id:
                sheet.delete_rows(cell.row, 1)
    
    wb.save(path)

    markup = tl.types.ReplyKeyboardMarkup()

    markup.row('Bachelor')
    markup.row('Master')

    bot.send_message(message.chat.id, 'Choose degree', reply_markup=markup)
    bot.register_next_step_handler(message,year_define)
    
def year_define(message):
    if message.text=='Bachelor':
        markup = tl.types.ReplyKeyboardMarkup()

        markup.row('2021 Bachelor')
        markup.row('2020 Bachelor')
        markup.row('2019 Bachelor')
        markup.row('2018 Bachelor')
    elif message.text=='Master':
        markup = tl.types.ReplyKeyboardMarkup()

        markup.row('2021 Master')


    bot.send_message(message.chat.id, 'Choose Admission year', reply_markup=markup)
    bot.register_next_step_handler(message,group_define)

def group_define(message):
    if message.text=='2021 Bachelor':
        markup = tl.types.ReplyKeyboardMarkup()

        markup.row('A 1','A 2')
        markup.row('B 1','B 2')
        markup.row('C 1','C 2')
        markup.row('D 1','D 2')
    elif message.text=='2020 Bachelor':
        markup = tl.types.ReplyKeyboardMarkup()

        markup.row('CE-20 1','CE-20 2')
        markup.row('CS-20 1','CS-20 2')
        markup.row('GE-20 1','GE-20 2')
        markup.row('PE-20 1','PE-20 2')
    elif message.text=='2019 Bachelor':
        markup = tl.types.ReplyKeyboardMarkup()

        markup.row('CE-19 1','CE-19 2')
        markup.row('CS-19 1','CS-19 2')
        markup.row('GE-19 1','GE-19 2')
        markup.row('PE-19 1','PE-19 2')
    elif message.text=='2018 Bachelor':
        markup = tl.types.ReplyKeyboardMarkup()

        markup.row('CE-18 1','CE-18 2')
        markup.row('CS-18 1','CS-18 2')
        markup.row('GE-18 1','GE-18 2')
        markup.row('PE-18 1','PE-18 2')
    elif message.text=='2021 Master':
        markup = tl.types.ReplyKeyboardMarkup()

        markup.row('M1 DSAI-21 1','M1 DSAI-21 2')
        markup.row('M1 GE-21 1','M1 GE-21 2')
        markup.row('M1 CE-21 1','M1 CE-21 2')

    bot.send_message(message.chat.id, 'Choose your group', reply_markup=markup)
    bot.register_next_step_handler(message,done)

def done(message):
    global path

    id=str(message.chat.id)
    student_group = message.text

    update_sheet(student_group,id,path)

    bot.send_message(id,f"Done! Your group is {student_group}")


'''CURRENT command to get current lesson'''
@bot.message_handler(commands=['current'])
def current(message):
    global path, current_day,current_week_number

    id = str(message.chat.id)
    student_group=get_student_group(id)

    if student_group=='':
        bot.send_message(id,"Firstly, /setgroup")
    else:
        g="Group "+student_group[-1]
        cLesson = _current_index(current_day,current_week_number,student_group[:-2],g)

        try:
            bot.send_message(message.chat.id, "Current lesson info: \n ~~ "+cLesson.name+" ~~ \nClassroom: "+cLesson.room+"\nTeacher: "+cLesson.teacher_name)
        except:
            bot.send_message(message.chat.id, "NO LESSON")
    

'''NEXT command to get next lesson'''
@bot.message_handler(commands=['next'])
def next(message):
    global path, current_day,current_week_number
    groups = get_groups_dict_from(path)
    id = str(message.chat.id)

    student_group=get_student_group(id)
    if student_group=='':
        bot.send_message(id,"Firstly, /setgroup")
    else:
        g="Group "+student_group[-1]
        nLesson = _next_index(current_day,current_week_number,student_group[:-2],g)

        try:
            bot.send_message(message.chat.id, "Next lesson info: \n ~~ "+nLesson.name+" ~~ \nClassroom: "+nLesson.room+"\nTeacher: "+nLesson.teacher_name)
        except:
            bot.send_message(message.chat.id, "NO LESSON")


'''TOMORROW command to get lessons for tomorrow'''
@bot.message_handler(commands=['tomorrow'])
def tomorrow(message):
    global path, current_day,current_week_number,tomorrow_day
    id = str(message.chat.id)
    student_group=get_student_group(id)
    if student_group=='':
        bot.send_message(id,"Firstly, /setgroup")
    else:
        try:
            g="Group "+student_group[-1]
            tLessons = _tomorrow_index(current_day,current_week_number,student_group[:-2],g)
            
            bot.send_message(id,"Lessons for "+str(tomorrows_date))

            if tLessons=="N":
                bot.send_message(id,"NO LESSONS")
            else:
                for i in range(-1,-1-len(tLessons),-1):
                    if tLessons[i]=="l":
                        del tLessons[i]
                    else:
                        break
                
                for lesson in tLessons:
                    try:
                        bot.send_message(id,"Lesson "+lesson.period+"  info:\n\n ~~ "+lesson.name+" ~~ \nClassroom: "+lesson.room+"\nTeacher: "+lesson.teacher_name)
                    except:
                        bot.send_message(id,"NO "+lesson.period+" LESSON")
        except:
            pass

'''TODAY command to get lessons for today'''
@bot.message_handler(commands=['today'])
def today(message):

    global path, current_day,current_week_number,todays_date
    id = str(message.chat.id)
    student_group=get_student_group(id)

    
    if student_group=='':
        bot.send_message(id,"Firstly, /setgroup")
        
    else:
        try:
            g="Group "+student_group[-1]
            tLessons = _today_index(current_day,current_week_number,student_group[:-2],g)
            
            bot.send_message(id,"Lessons for "+str(todays_date))

            if tLessons=="N":
                bot.send_message(id,"NO LESSONS")
            else:
                for i in range(-1,-1-len(tLessons),-1):
                    if tLessons[i]=="l":
                        del tLessons[i]
                    else:
                        break
                
                for lesson in tLessons:
                    try:
                        bot.send_message(id,"Lesson "+lesson.period+"  info:\n\n ~~ "+lesson.name+" ~~ \nClassroom: "+lesson.room+"\nTeacher: "+lesson.teacher_name)
                    except:
                        bot.send_message(id,"NO "+lesson.period+" LESSON")
        except:
            pass

bot.polling()
